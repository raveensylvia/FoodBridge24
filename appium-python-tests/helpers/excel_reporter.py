# ============================================================
# Excel Report Generator — Appium Tests (Python)
# ============================================================
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from config.appium_config import CONFIG

COLORS = {
    "PASS": "4CAF50",
    "FAIL": "F44336",
    "SKIP": "FFC107",
    "HEADER_BG": "0D3349",
    "HEADER_FG": "FFFFFF",
    "ALT_ROW": "F5F5F5",
    "TITLE_BG": "0A2740"
}

FONT_NAME = "Calibri"
FONT_HEADER = Font(name=FONT_NAME, size=11, bold=True, color=COLORS["HEADER_FG"])
FONT_TITLE = Font(name=FONT_NAME, size=16, bold=True, color=COLORS["HEADER_FG"])
FONT_PASS = Font(name=FONT_NAME, size=10, bold=True, color="2E7D32")
FONT_FAIL = Font(name=FONT_NAME, size=10, bold=True, color="C62828")
FONT_BOLD = Font(name=FONT_NAME, size=10, bold=True)
FONT_BODY = Font(name=FONT_NAME, size=10)

def generate_report(results, summary):
    report_dir = CONFIG["report_dir"]
    if not os.path.exists(report_dir):
        os.makedirs(report_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    filename = f"{CONFIG['report_name']}_{timestamp}.xlsx"
    filepath = os.path.join(report_dir, filename)

    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets
    _build_summary_sheet(wb, summary, results)
    _build_details_sheet(wb, results)
    _build_category_sheet(wb, results)

    wb.save(filepath)
    print(f"\n  [Report] Mobile Excel report saved -> {filepath}\n")
    return filepath

def _build_summary_sheet(wb, summary, results):
    ws = wb.create_sheet(title="📋 Summary")
    ws.sheet_properties.tabColor = COLORS["HEADER_BG"]

    # Set Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "📱 FoodBridge Android App — Test Report"
    title_cell.font = FONT_TITLE
    title_cell.fill = PatternFill(start_color=COLORS["TITLE_BG"], end_color=COLORS["TITLE_BG"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Metadata
    meta = [
        ("Generated At", datetime.now().strftime("%d/%m/%Y, %H:%M:%S")),
        ("Platform", "Android"),
        ("Device", CONFIG["capabilities"]["appium:deviceName"]),
        ("Android Version", CONFIG["capabilities"]["appium:platformVersion"]),
        ("Total Duration", f"{(summary['duration'] / 1000):.2f}s")
    ]
    for idx, (k, v) in enumerate(meta):
        row_num = idx + 2
        ws.cell(row=row_num, column=1, value=k).font = FONT_BOLD
        ws.cell(row=row_num, column=2, value=v).font = FONT_BODY
        ws.row_dimensions[row_num].height = 18

    # Stats cards
    stats = [
        {"label": "Total Tests", "value": summary["total"], "color": "0D3349"},
        {"label": "✅ Passed", "value": summary["passed"], "color": "4CAF50"},
        {"label": "❌ Failed", "value": summary["failed"], "color": "F44336"},
        {"label": "⏭ Skipped", "value": summary["skipped"], "color": "FFC107"}
    ]
    
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 20

    for col_idx, s in enumerate(stats, start=1):
        ws.merge_cells(start_row=8, start_column=col_idx, end_row=9, end_column=col_idx)
        cell = ws.cell(row=8, column=col_idx)
        cell.value = f"{s['label']}\n{s['value']}"
        cell.font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=s["color"], end_color=s["color"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    pass_rate = (summary["passed"] / summary["total"] * 100) if summary["total"] > 0 else 0
    ws.cell(row=11, column=1, value=f"Pass Rate: {pass_rate:.1f}%").font = Font(name=FONT_NAME, size=12, bold=True, color="2E7D32")
    ws.row_dimensions[11].height = 20

    # Category breakdown table header
    headers = ["Category", "Total", "Passed", "Failed", "Pass %"]
    header_row = 13
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color=COLORS["HEADER_BG"], end_color=COLORS["HEADER_BG"], fill_type="solid")
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[header_row].height = 20

    # Calculate categories
    by_category = {}
    for r in results:
        cat = r["category"]
        if cat not in by_category:
            by_category[cat] = {"total": 0, "passed": 0, "failed": 0}
        by_category[cat]["total"] += 1
        if r["status"] == "PASS":
            by_category[cat]["passed"] += 1
        elif r["status"] == "FAIL":
            by_category[cat]["failed"] += 1

    current_row = 14
    for idx, (cat, data) in enumerate(by_category.items()):
        pct = (data["passed"] / data["total"] * 100) if data["total"] > 0 else 0
        row_values = [cat, data["total"], data["passed"], data["failed"], f"{pct:.0f}%"]
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = FONT_BODY
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
            
            # Alt row shading
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLORS["ALT_ROW"], end_color=COLORS["ALT_ROW"], fill_type="solid")
            
            if col_idx == 3:
                cell.font = FONT_PASS
            elif col_idx == 4:
                cell.font = FONT_FAIL
        ws.row_dimensions[current_row].height = 18
        current_row += 1

def _build_details_sheet(wb, results):
    ws = wb.create_sheet(title="📝 Test Details")
    ws.sheet_properties.tabColor = "4CAF50"

    headers = [
        ("Test ID", 10),
        ("Category", 25),
        ("Test Name", 50),
        ("Type", 18),
        ("Status", 12),
        ("Duration (s)", 15),
        ("Error / Notes", 55)
    ]
    
    # Header Row
    ws.row_dimensions[1].height = 24
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color=COLORS["HEADER_BG"], end_color=COLORS["HEADER_BG"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = w

    # Rows
    thin_border = Border(
        left=Side(style='thin', color="DDDDDD"),
        right=Side(style='thin', color="DDDDDD"),
        top=Side(style='thin', color="DDDDDD"),
        bottom=Side(style='thin', color="DDDDDD")
    )

    for idx, r in enumerate(results):
        row_num = idx + 2
        dur_val = f"{(r['duration'] / 1000):.2f}" if r["duration"] != "-" and isinstance(r["duration"], (int, float)) else "-"
        
        ws.cell(row=row_num, column=1, value=r["id"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=r["category"])
        ws.cell(row=row_num, column=3, value=r["name"])
        ws.cell(row=row_num, column=4, value=r.get("type", "Functional"))
        
        sc = ws.cell(row=row_num, column=5, value=r["status"])
        sc.alignment = Alignment(horizontal="center")
        if r["status"] == "PASS":
            sc.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
            sc.font = FONT_PASS
        elif r["status"] == "FAIL":
            sc.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            sc.font = FONT_FAIL
        else:
            sc.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            sc.font = FONT_BOLD

        ws.cell(row=row_num, column=6, value=dur_val).alignment = Alignment(horizontal="center")
        
        err_cell = ws.cell(row=row_num, column=7, value=r.get("error", ""))
        err_cell.alignment = Alignment(wrap_text=True)

        for col_idx in range(1, 8):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if cell.font == FONT_BODY or cell.font is None:
                cell.font = FONT_BODY
            
            # Alt row shading
            if idx % 2 == 0 and r["status"] != "PASS" and r["status"] != "FAIL":
                cell.fill = PatternFill(start_color=COLORS["ALT_ROW"], end_color=COLORS["ALT_ROW"], fill_type="solid")
            elif idx % 2 == 0 and col_idx != 5:
                cell.fill = PatternFill(start_color=COLORS["ALT_ROW"], end_color=COLORS["ALT_ROW"], fill_type="solid")

        ws.row_dimensions[row_num].height = 20

    ws.auto_filter.ref = f"A1:G{len(results) + 1}"
    ws.freeze_panes = "A2"

def _build_category_sheet(wb, results):
    ws = wb.create_sheet(title="📊 By Category")
    ws.sheet_properties.tabColor = "E94560"

    headers = [
        ("Category", 30),
        ("Test Type", 20),
        ("Test Name", 50),
        ("Status", 12)
    ]
    
    # Header Row
    ws.row_dimensions[1].height = 22
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color=COLORS["HEADER_BG"], end_color=COLORS["HEADER_BG"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = w

    # Sorted Results
    sorted_results = sorted(results, key=lambda x: (x["category"], x.get("type", "Functional"), x["id"]))

    for idx, r in enumerate(sorted_results):
        row_num = idx + 2
        ws.cell(row=row_num, column=1, value=r["category"])
        ws.cell(row=row_num, column=2, value=r.get("type", "Functional"))
        ws.cell(row=row_num, column=3, value=r["name"])
        
        sc = ws.cell(row=row_num, column=4, value=r["status"])
        sc.alignment = Alignment(horizontal="center")
        if r["status"] == "PASS":
            sc.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
            sc.font = FONT_PASS
        elif r["status"] == "FAIL":
            sc.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            sc.font = FONT_FAIL

        for col_idx in range(1, 5):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.font == FONT_BODY or cell.font is None:
                cell.font = FONT_BODY
            
            # Alt row shading
            if idx % 2 == 0 and col_idx != 4:
                cell.fill = PatternFill(start_color=COLORS["ALT_ROW"], end_color=COLORS["ALT_ROW"], fill_type="solid")
        ws.row_dimensions[row_num].height = 18

    ws.auto_filter.ref = f"A1:D{len(results) + 1}"
    ws.freeze_panes = "A2"
