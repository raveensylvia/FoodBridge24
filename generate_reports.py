import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def load_json(filepath):
    if not os.path.exists(filepath):
        print(f"Warning: {filepath} not found.")
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return []

def create_excel_report(data, filename, report_title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    pass_font = Font(color="008000", bold=True)
    fail_font = Font(color="FF0000", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Report Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = report_title
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = center_align

    # Headers
    headers = ["Test ID", "Category", "Test Name", "Type", "Status", "Duration (ms)", "Error Message"]
    ws.append([]) # Empty row for spacing
    ws.append(headers)

    header_row = ws[3]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data Rows
    total, passed, failed = 0, 0, 0
    for item in data:
        status = item.get("status", "FAIL")
        total += 1
        if status == "PASS":
            passed += 1
        else:
            failed += 1

        row = [
            item.get("id", ""),
            item.get("category", ""),
            item.get("name", ""),
            item.get("type", ""),
            status,
            item.get("duration", 0),
            item.get("error", "")
        ]
        ws.append(row)
        
        current_row = ws[ws.max_row]
        for idx, cell in enumerate(current_row):
            cell.border = thin_border
            if idx == 4: # Status column
                cell.font = pass_font if status == "PASS" else fail_font
                cell.alignment = center_align
            elif idx in (0, 3, 5):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-adjust column widths
    column_widths = [12, 25, 60, 15, 12, 15, 50]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = column_width

    # Summary Sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    
    summary_data = [
        ("Report Title", report_title),
        ("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Tests", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Pass Rate", f"{(passed/total)*100:.2f}%" if total > 0 else "0.00%")
    ]
    
    for row in summary_data:
        ws_summary.append(row)
        
    for r in ws_summary.iter_rows(min_row=1, max_row=len(summary_data)):
        r[0].font = Font(bold=True)
        r[0].border = thin_border
        r[1].border = thin_border

    # Ensure output directory exists
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)
    print(f"Generated report: {filename} ({total} tests)")


def main():
    web_file = os.path.join(".web_results.json")
    mobile_file = os.path.join("appium-python-tests", ".pytest_results.json")
    
    web_data = load_json(web_file)
    mobile_data = load_json(mobile_file)
    
    # 1. Selenium Report (Exclude Vulnerability)
    selenium_tests = [d for d in web_data if "Vulnerability" not in d.get("category", "")]
    if selenium_tests:
        create_excel_report(
            selenium_tests, 
            os.path.join("reports", "Selenium_Web_Test_Report.xlsx"), 
            "Selenium Web Functional & UI Test Report"
        )
        
    # 2. Appium Report (Exclude Vulnerability)
    appium_tests = [d for d in mobile_data if "Vulnerability" not in d.get("category", "")]
    if appium_tests:
        create_excel_report(
            appium_tests, 
            os.path.join("reports", "Appium_Mobile_Test_Report.xlsx"), 
            "Appium Mobile Functional & E2E Test Report"
        )
        
    # 3. Vulnerability Report (Only Vulnerability from Web and Mobile)
    vuln_tests = [d for d in web_data if "Vulnerability" in d.get("category", "")] + \
                 [d for d in mobile_data if "Vulnerability" in d.get("category", "")]
    if vuln_tests:
        create_excel_report(
            vuln_tests, 
            os.path.join("reports", "Security_Vulnerability_Test_Report.xlsx"), 
            "Comprehensive Security & Vulnerability Test Report"
        )

if __name__ == "__main__":
    main()
